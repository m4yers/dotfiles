#!/usr/bin/env python3
"""Read an .xlsx file and print contents as TSV."""

import sys
from openpyxl import load_workbook


def main():
    if len(sys.argv) < 2:
        print("Usage: read-xlsx.py <file.xlsx> [sheet_name]",
              file=sys.stderr)
        sys.exit(1)

    path = sys.argv[1]
    sheet_filter = sys.argv[2] if len(sys.argv) > 2 else None

    wb = load_workbook(path, data_only=True, read_only=True)

    for name in wb.sheetnames:
        if sheet_filter and name != sheet_filter:
            continue
        ws = wb[name]
        print("=== {} ===".format(name))
        for row in ws.iter_rows(values_only=True):
            print('\t'.join(
                str(c) if c is not None else '' for c in row
            ))

    wb.close()


if __name__ == '__main__':
    main()
